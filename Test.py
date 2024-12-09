# # def bill_export_excel(self):
# #     rows = []
# #     for rec in self:
# #         rows.append([rec.bill_number, rec.date, rec.user_id.user_name, rec.grand_total])
#
# #     # rows = [ [rec.bill_number, str(rec.date), rec.user_id.user_name, rec.grand_total] for rec in self]
#
#
# # # create file
# #     with open('bill_date.csv', mode='w') as file:
# #         writer = csv.writer(file, delimiter=',')
# #         writer.writerow(['Bill No', 'Date', 'User Name', 'Grand Total'])
# #         for row in rows:
# #             writer.writerow(row)
#
# #     # Read file and download
# #     with open('bill_date.csv', 'r', encoding='UTF-8') as f2:
# #         data = str.encode(f2.read(), 'UTF-8')
# #         wiz_id = self.env['retail.bill.excel.wizard'].create({
# #             'file': base64.encodebytes(data)
# #         })
# #         os.remove('bill_date.csv')
# #         return {
# #             'type': 'ir.actions.act_url',
# #             # 'url': 'web/content/?model=retail.bill.excel.wizard&download=true&field=file&id=%s&filename=%s.csv' % (wiz_id.id, 'bill_excel_date'),
# #             'url': 'web/content/?model=retail.bill.excel.wizard&download=true&field=file&id={}&filename={}.csv'.format(wiz_id.id, 'bill_excel_date'),
# #             'target': 'self',
# #         }
#
#
# import io
# from openpyxl import load_workbook
#
#
# def read_xlsx_to_dict(file_path):
#     # Open the file in binary mode
#     with open(file_path, 'rb') as f:
#         file_data = f.read()
#
#     # Load the workbook from the binary data
#     read_file = load_workbook(io.BytesIO(file_data))
#
#     # Get the name of the first sheet
#     all_sheet = read_file.sheetnames[0]
#
#     # Get the sheet object
#     sheet = read_file[all_sheet]
#
#     # Get column headers from the first row
#     headers = [cell.value for cell in sheet[1]]  # First row contains headers
#
#     # Prepare a list to store rows as dictionaries
#     data_dict_list = []
#
#     # Iterate over the rows, starting from the second row
#     for row in sheet.iter_rows(min_row=3, values_only=True):
#         row_dict = {}
#         for idx, cell in enumerate(row):
#             header = headers[idx]  # Use header as key
#             if header not in row_dict:
#                 row_dict[header] = cell
#
#                 # row_dict[header] = row_dict[header].append(cell)
#             # row_dict[header] =
#
#         data_dict_list.append(row_dict)
#
#     return data_dict_list
#
# # Example usage
# file_path = 'Template Concept for script.xlsx'  # Provide the correct path to your Excel file
# data = read_xlsx_to_dict(file_path)
#
# # Print the resulting list of dictionaries
# for row in data:
#     print(row)
#
#
# # default_code	Ignore	name	attribute_id	value_ids	attribute_id	value_ids	attribute_id	value_ids	attribute_id	value_ids	categ_id	uom_id	uom_po_id	weight	standard_price	lst_price	Description	Packaging	Packaging Contains	Packaging	Packaging Contains	Packaging	Packaging Contains	Price list	Price list	Price list	Vendors/Vendor	Vendors/Vendor Product Code	Vendors/Price
# # Internal Reference	Original Full Product Name	Name	Attribute 1	Attribute Values 1	Attribute 2	Attribute Values 2	Attribute 3	Attribute Values 3	Attribute 4	Attribute Values 4	Category	UNIT MEASURE	Purchase UOM	WEIGHT	Cost	Sales Price	Description	U/M LABEL ONE	U/M QTY ONE	U/M LABEL TWO	U/M QTY TWO	U/M LABEL THREE	U/M QTY THREE	Wholesale	Retail	MSRP	Vendors/Vendor	Vendors/Vendor Product Code	Vendors/Price
# # ESXL3DE	Emerson Smooth XL Desert 114sq	Emerson XL	LIBERTY TEXTURE	Smooth	LIBERTY COLORS	Desert					All / Pavers	Units			28.842	446.88	Daron LibertyStone 30/35%mu  10.36sq/layer  330.42	CUBES	1	LAYERS	0.091	PIECES	1						28.842
# # ESXL3WM	Emerson Smooth XL West Mountain 114sq	Emerson XL	LIBERTY TEXTURE	Smooth	LIBERTY COLORS	West Mountain					All / Pavers	Units			28.842	446.88	Daron 30-35%mu 114 sq,ft 30.36sq,ft/layer  $330.42	CUBES	1	LAYERS	0.091	PIECES	1						28.842
# # ESXL3WI	Emerson Smooth XL Willow 114sq	Emerson XL	LIBERTY TEXTURE	Smooth	LIBERTY COLORS	Willow					All / Pavers	Units			28.842	446.88	Daron LibertyStone 30/35%mu  10.36sq/layer  330.42	CUBES	1	LAYERS	0.091	PIECES	1						28.842
# # EXL3DE	Emerson XL Desert 114sq	Emerson XL	LIBERTY TEXTURE	Slate	LIBERTY COLORS	Desert					All / Pavers	Units			28.842	446.88	Daron LibertyStone 30/35%mu  10.36sq/layer  330.42	CUBES	1	LAYERS	0.091	PIECES	1						28.842
# # EXL3WM	Emerson XL West Mountain 114sq	Emerson XL	LIBERTY TEXTURE	Slate	LIBERTY COLORS	West Mountain					All / Pavers	Units			28.842	446.88	Daron LibertyStone 30/35%mu  10.36sq/layer  330.42	CUBES	1	LAYERS	0.091	PIECES	1						28.842
# # EXL3WI	Emerson XL Willow 114sq	Emerson XL	LIBERTY TEXTURE	Slate	LIBERTY COLORS	Willow					All / Pavers	Units			28.842	446.88	Daron LibertyStone 30/35%mu  10.36sq/layer  330.42	CUBES	1	LAYERS	0.091	PIECES	1						28.842
# #
# # for this xlsx file data to prepper muti row add same key in dict. in python
#
# # [{
# #     'default_code': 'ESXL3DE',
# #     'Ignore': 'Emerson Smooth XL Desert 114sq',
# #     'name': 'Emerson XL',
# #     'attribute_id': ['LIBERTY TEXTURE','LIBERTY COLORS'] ,
# #     'value_ids': [['Smooth'],['Desert']],
# #     'categ_id': 'All / Pavers',
# #     'uom_id': 'Units',
# #     'uom_po_id': '',
# #     'weight': 28.842,
# #     'standard_price': 446.88,
# #     'lst_price': 330.42,
# #     'Description': 'Daron LibertyStone 30/35%mu  10.36sq/layer  330.42',
# #     'Packaging': ['CUBES',]
# #     'Packaging Contains': [1,]
# #     'Price list': 'LAYERS',
# #     'Vendors/Vendor': 'PIECES',
# #     'Vendors/Vendor Product Code': 1,
# #     'Vendors/Price': 28.842
# # },
# # {
# #     'default_code': 'ESXL3WM',
# #     'Ignore': 'Emerson Smooth XL West Mountain 114sq',
# #     'name': 'Emerson XL',
# #     'attribute_id': 'LIBERTY TEXTURE',
# #     'value_ids': 'Smooth',
# #     'attribute_id': 'LIBERTY COLORS',
# #     'value_ids': 'West Mountain',
# #     'attribute_id': '',
# #     'value_ids': '',
# #     'attribute_id': '',
# #     'value_ids': '',
# #     'categ_id': 'All / Pavers',
# #     'uom_id': 'Units',
# #     'uom_po_id': '',
# #     'weight': 28.842,
# #     'standard_price': 446.88,
# #     'lst_price': 330.42,
# #     'Description': 'Daron 30-35%mu 114 sq,ft 30.36sq,ft/layer  $330.42',
# #     'Packaging': 'CUBES',
# #     'Packaging Contains': 1,
# #     'Price list': 'LAYERS',
# #     'Vendors/Vendor': 'PIECES',
# #     'Vendors/Vendor Product Code': 1,
# #     'Vendors/Price': 28.842
# # }]-
#
# [
#   {
#     "default_code": "ESXL3DE",
#     "Ignore": "Emerson Smooth XL Desert 114sq",
#     "name": "Emerson XL",
#     'attribute_id': ['LIBERTY TEXTURE','LIBERTY COLORS'] ,
#     'value_ids': [['Smooth'],['Desert','red']],
#     "categ_id": "All / Pavers",
#     "uom_id": "Units",
#     "uom_po_id": null,
#     "weight": null,
#     "standard_price": 28.842,
#     "lst_price": 446.88,
#     "Description": "Daron LibertyStone 30/35%mu  10.36sq/layer  330.42",
#     "Packaging":[ "CUBES",'LAYERS',"PIECES"],
#     "Packaging Contains": [1,0.091,1],
#     "Price list": [],
#     "Vendors/Vendor": null,
#     "Vendors/Vendor Product Code": null,
#     "Vendors/Price": 28.842
#   },
#   {
#     "default_code": "ESXL3WM",
#     "Ignore": "Emerson Smooth XL West Mountain 114sq",
#     "name": "Emerson XL",
#     'attribute_id': ['LIBERTY TEXTURE','LIBERTY COLORS'] ,
#     'value_ids': [['Smooth'],['West Mountain']],
#     "categ_id": "All / Pavers",
#     "uom_id": "Units",
#     "uom_po_id": null,
#     "weight": null,
#     "standard_price": 28.842,
#     "lst_price": 446.88,
#     "Description": "Daron 30-35%mu 114 sq,ft 30.36sq,ft/layer  $330.42",
#     "Packaging":[ "CUBES",'LAYERS',"PIECES"],
#     "Packaging Contains": [1,0.091,1],
#     "Price list": [],
#     "Vendors/Vendor": null,
#     "Vendors/Vendor Product Code": null,
#     "Vendors/Price": 28.842
#   },
#   {
#     "default_code": "ESXL3WI",
#     "Ignore": "Emerson Smooth XL Willow 114sq",
#     "name": "Emerson XL",
#     'attribute_id': ['LIBERTY TEXTURE','LIBERTY COLORS'] ,
#     'value_ids': [['Smooth'],['Willow']],
#     "categ_id": "All / Pavers",
#     "uom_id": "Units",
#     "uom_po_id": null,
#     "weight": null,
#     "standard_price": 28.842,
#     "lst_price": 446.88,
#     "Description": "Daron LibertyStone 30/35%mu  10.36sq/layer  330.42",
#     "Packaging":[ "CUBES",'LAYERS',"PIECES"],
#     "Packaging Contains": [1,0.091,1],
#     "Price list": [],
#     "Vendors/Vendor": null,
#     "Vendors/Vendor Product Code": null,
#     "Vendors/Price": 28.842
#   },
#   {
#     "default_code": "EXL3DE",
#     "Ignore": "Emerson XL Desert 114sq",
#     "name": "Emerson XL",
#     'attribute_id': ['LIBERTY TEXTURE','LIBERTY COLORS'] ,
#     'value_ids': [['Slate'],['Desert']],
#     "categ_id": "All / Pavers",
#     "uom_id": "Units",
#     "uom_po_id": null,
#     "weight": null,
#     "standard_price": 28.842,
#     "lst_price": 446.88,
#     "Description": "Daron LibertyStone 30/35%mu  10.36sq/layer  330.42",
#     "Packaging":[ "CUBES",'LAYERS',"PIECES"],
#     "Packaging Contains": [1,0.091,1],
#     "Price list": [],
#     "Vendors/Vendor": null,
#     "Vendors/Vendor Product Code": null,
#     "Vendors/Price": 28.842
#   },
#   {
#     "default_code": "EXL3WM",
#     "Ignore": "Emerson XL West Mountain 114sq",
#     "name": "Emerson XL",
#     'attribute_id': ['LIBERTY TEXTURE','LIBERTY COLORS'],
#     'value_ids': [['Slate'],['West Mountain']],
#     "categ_id": "All / Pavers",
#     "uom_id": "Units",
#     "uom_po_id": null,
#     "weight": null,
#     "standard_price": 28.842,
#     "lst_price": 446.88,
#     "Description": "Daron LibertyStone 30/35%mu  10.36sq/layer  330.42",
#     "Packaging":[ "CUBES",'LAYERS',"PIECES"],
#     "Packaging Contains": [1,0.091,1],
#     "Price list": [],
#     "Vendors/Vendor": null,
#     "Vendors/Vendor Product Code": null,
#     "Vendors/Price": 28.842
#   },
#   {
#     "default_code": "EXL3WI",
#     "Ignore": "Emerson XL Willow 114sq",
#     "name": "Emerson XL",
#     'attribute_id': ['LIBERTY TEXTURE','LIBERTY COLORS'],
#     'value_ids': [['Slate'],['Willow']],
#     "categ_id": "All / Pavers",
#     "uom_id": "Units",
#     "uom_po_id": null,
#     "weight": null,
#     "standard_price": 28.842,
#     "lst_price": 446.88,
#     "Description": "Daron LibertyStone 30/35%mu  10.36sq/layer  330.42",
#     "Packaging":[ "CUBES",'LAYERS',"PIECES"],
#     "Packaging Contains": [1,0.091,1],
#     "Price list": [],
#     "Vendors/Vendor": null,
#     "Vendors/Vendor Product Code": null,
#     "Vendors/Price": 28.842
#   }
# ]
#
#

import io
from openpyxl import load_workbook

def read_xlsx_to_dict(file_path):
    # Open the file in binary mode
    with open(file_path, 'rb') as f:
        file_data = f.read()

    # Load the workbook from the binary data
    read_file = load_workbook(io.BytesIO(file_data))

    # Get the name of the first sheet
    all_sheet = read_file.sheetnames[0]

    # Get the sheet object
    sheet = read_file[all_sheet]

    # Get column headers from the first row
    headers = [cell.value for cell in sheet[1]]  # First row contains headers

    # Prepare a list to store rows as dictionaries
    data_dict_list = []

    # Iterate over the rows, starting from the second row (adjust based on your data's row structure)
    for row in sheet.iter_rows(min_row=3, values_only=True):
        row_dict = {}
        attribute_id = []
        value_ids = []
        packaging = []
        packaging_contains = []

        # Iterate over each cell in the row
        for idx, cell in enumerate(row):
            header = headers[idx]

            # Collect 'attribute_id' and 'value_ids' in lists
            if 'attribute_id' == header:
                attribute_id.append(cell) if cell else None
            elif 'value_ids' == header:
                value_ids.append(cell.split(',')) if cell else None  # Wrap it in a list as requested

            # Collect 'Packaging' and 'Packaging Contains' in lists
            elif 'Packaging' == header:
                packaging.append(cell)
            elif 'Packaging Contains' == header:
                packaging_contains.append(cell)

            # For other keys, directly add the value to the row_dict
            elif header not in row_dict:
                row_dict[header] = cell

        # Now assign lists to their respective fields in the row_dict
        row_dict['attribute_id'] = attribute_id
        row_dict['value_ids'] = value_ids
        row_dict['Packaging'] = packaging
        row_dict['Packaging Contains'] = packaging_contains

        # Add the row dictionary to the list of data
        data_dict_list.append(row_dict)

    return data_dict_list

# Example usage
file_path = 'Template Concept for script.xlsx'  # Provide the correct path to your Excel file
data = read_xlsx_to_dict(file_path)

# Print the resulting list of dictionaries
# import json
# print(json.dumps(data, indent=2))
# #
print(data)
